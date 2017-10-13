from VotingApp.mainApp import app, db, update_ret, add_stock, update_score, create_stock_info
from VotingApp.models import User, Tickers, Role, Transactions, Stock
from flask_script import Manager, prompt_bool
from openpyxl import load_workbook, Workbook
import requests

manager = Manager(app)

@manager.command
def initdb():
    db.create_all()
    admin = Role(name="Admin", description="Gets to look at all the rankings.")
    db.session.add(User(firstName="Arnav", lastName="Jain",email="arnav@utexas.edu", password="test11", roles=[admin]))
    db.session.commit()
    refreshdb()
    print 'Initialized the database'

@manager.command
def dropdb():
    if prompt_bool(
        "Are you sure you want to lose all your data"):
        db.drop_all()
        print 'Dropped the database'

@manager.command
def print_stocks():
    for stock in Stock.query.all():
        print(str(stock.id) + ". Stock: " + stock.ticker + ", " + str(stock.price))

    for ticker in Tickers.query.all():
        print(str(ticker.id) + ". Ticker: " + ticker.ticker + ", " + str(ticker.startingPrice) + ", " + str(ticker.short))

    for user in User.query.all():
        for stock in user.stocks:
            print(str(stock.id) + ". " + user.email + " voted for " + stock.ticker)
        for transaction in user.transactions:
            print(str(transaction.id) + ". " +  user.email + " took a position on " + transaction.ticker + " with return of " + str(transaction.returns))

@manager.command
def print_returns():
    for user in User.query.all():
        print(str(user.email) + " has a return of " + str(user.ret) + "%.")

@manager.command
def delete_transactions():
    continue_or_not = 'y'

    while continue_or_not == 'y':
        id = int(input("What is the ID? "))
        transaction = Transactions.query.filter_by(id = id).first()
        db.session.delete(transaction)
        db.session.commit()
        continue_or_not = str(raw_input("Do you want to continue (y/n)? ")).lower()

@manager.command
def refreshdb():
    # Refresh the score and ranks for each student
    for student in User.query.all():
        numStocks = update_ret(student, student.stocks, student.transactions)
        update_score(student, student.ret, numStocks)

    # Refresh the stored information for each stock
    for stock in Tickers.query.all():
        create_stock_info(stock)
    for stock in Transactions.query.all():
        create_stock_info(stock)

@manager.command
def get_user_emails():
    for user in User.query.all():
        print(user.email)

@manager.command
def addstock():
    name = str(raw_input("What is the file name? "))
    symbol = str(raw_input("Ticker: "))
    price = int(raw_input("Starting price: $"))
    num_votes = int(raw_input("How many votes were there? "))

    wb = load_workbook(filename=name)
    ws = wb.active

    for index in range(1, num_votes + 1):
        if str(ws['B' + str(index)].value) == 'Long':
            if  Tickers.query.filter_by(short = False, ticker = symbol).count() > 0:
                stock = Tickers.query.filter_by(short = False, ticker = symbol).first()
                print("I FOUND THE STOCK ALREADY")
            else:
                stock = Tickers(ticker=symbol, startingPrice=price, short=False)
        elif str(ws['B' + str(index)].value) == 'Short':
            if Tickers.query.filter_by(short = True, ticker = symbol).count() > 0:
                stock = Tickers.query.filter_by(short = True, ticker = symbol).first()
                print("I FOUND THE STOCK ALREADY")
            else:
                stock = Tickers(ticker=symbol, startingPrice=price, short=True)
        if User.query.filter_by(email=str(ws['A'+str(index)].value).lower()) != None:
            student = User.query.filter_by(email=str(ws['A'+str(index)].value)).first()
            add_stock(student, stock)

@manager.command
def fixdb():
    num_students = User.query.count()
    num_students_wo_id = 1
    for student in User.query.all():
        if student.id == None:
            student.id == num_students + num_students_wo_id
            num_students_wo_id += 1
            print("I fixed", student.firstName, "!")

        db.session.commit()

if __name__ == '__main__':
    manager.run()


